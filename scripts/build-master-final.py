# Build PRODUCT-MASTER-FINAL.xlsx:
#   base = PRODUCT-MASTERb2.xlsx (latest core 245-SKU master)
#   + add/refresh sheet "สินค้าส่งด่วน (Express)" reconciled from the real-photo project
# Merges: express-products.json (supplier cost/moq/lead) + products-raw.json (colors)
#         + reconcile.json (new real-photo work) + featured-skus.json (live-on-web set)
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
def J(p): return json.load(open(os.path.join(ROOT, p), encoding="utf-8"))

exprod = {p["sku"]: p for p in J("scripts/express-products.json")}
raw = {p["sku"]: p for p in J("src/data/products-raw.json") if p["sku"].startswith("EX")}
recon = {e["sku"]: e for e in J("express-realphoto-2026/reconcile.json")}
featured = set(J("express-realphoto-2026/featured-skus.json"))

NAVY = "13244A"; GOLD = "F4B223"; CLOUD = "EEF1F5"
def exnum(s): return int(s[2:]) if s[2:].isdigit() else 9999
skus = sorted(raw.keys(), key=exnum)

wb = openpyxl.load_workbook(os.path.join(ROOT, "PRODUCT-MASTERb2.xlsx"))
title = "สินค้าส่งด่วน (Express)"
if title in wb.sheetnames:
    del wb[title]
ws = wb.create_sheet(title)

headers = ["SKU","ซัพพลายเออร์","ชื่อสินค้า","หมวดหมู่","จำนวนสี","สีทั้งหมด",
           "รูปจริง","ไฟล์รูปหลัก","บนเว็บ Express","MOQ","Lead time",
           "ต้นทุน/ชิ้น (฿)","วิธีสกรีน","ที่มา","สถานะ","หมายเหตุ"]
ws.append(headers)

for sku in skus:
    p = raw[sku]; ep = exprod.get(sku, {}); rc = recon.get(sku, {})
    colors = p.get("colors") or []
    has_photo = sku in featured or bool(p.get("colors"))
    # image presence: real photo if in reconcile (new studio) OR has gallery via featured
    real = "✓ ใหม่ (สตูดิโอ)" if sku in recon else ("✓" if sku in featured else "—")
    hero = (rc.get("hero_file") or "")
    on_web = "✓ Live" if sku in featured else "—"
    cost = ""
    if ep:
        cmin, cmax = ep.get("cost_min_thb"), ep.get("cost_max_thb")
        if cmin and cmax: cost = f"{cmin}-{cmax}" if cmin!=cmax else f"{cmin}"
        elif cmin: cost = str(cmin)
    lead = ep.get("lead_time_raw") or p.get("lead_time") or "7-14 วัน"
    sup = (ep.get("sup_name") or "") + ((" ("+ep.get("sup_code")+")") if ep.get("sup_code") else "")
    method = ep.get("custom_method") or (",".join(p.get("free_logo") or []))
    origin = "ใหม่ (รูปจริง 2026-06-20)" if rc.get("is_new") else ("รีโฟโต้สตูดิโอ" if sku in recon else "เดิม")
    status = []
    if rc.get("hero_is_group"): status.append("รูปรวม/กลุ่ม (ควรหาภาพเดี่ยว)")
    if not has_photo: status.append("ยังไม่มีรูป")
    if not colors: status.append("ยังไม่มีลิสต์สี")
    status = " · ".join(status) if status else "พร้อม"
    note = rc.get("notes","") if rc else ""
    ws.append([sku, sup, p.get("name",""), p.get("category_slug",""), len(colors),
               " / ".join(colors), real, hero, on_web, p.get("moq",""), lead,
               cost, method, origin, status, note[:240]])

# ---- styling ----
thin = Side(style="thin", color="D9DEE6")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
hfill = PatternFill("solid", fgColor=NAVY)
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = hfill; cell.font = Font(bold=True, color="FFFFFF", size=11, name="Tahoma")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
widths = [9,26,34,12,8,46,15,26,13,7,12,12,18,22,28,46]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for r in range(2, ws.max_row+1):
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=r, column=c); cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=(c in (6,16)))
        cell.font = Font(size=10, name="Tahoma")
    if r % 2 == 0:
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=CLOUD)
    # highlight NEW skus
    if recon.get(ws.cell(row=r,column=1).value, {}).get("is_new"):
        ws.cell(row=r, column=1).font = Font(size=10, bold=True, color="0A7A30", name="Tahoma")
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
ws.sheet_view.showGridLines = False

# move Express tab right after Master
order = wb.sheetnames
wb.move_sheet(title, offset=-(len(order)-1 - order.index("รายการสินค้า (Master)") - 1))

out = os.path.join(ROOT, "PRODUCT-MASTER-FINAL.xlsx")
wb.save(out)
live = sum(1 for s in skus if s in featured)
newc = sum(1 for s in skus if recon.get(s,{}).get("is_new"))
print(f"wrote {out}")
print(f"Express rows: {len(skus)} | live-on-web: {live} | NEW real-photo: {newc}")
print("sheets:", wb.sheetnames)
